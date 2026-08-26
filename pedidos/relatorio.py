"""
Gera o relatório de empréstimos em .xlsx no formato usado pela equipe
(planilha modelo com abas "Resumo" e "Pedidos"). Alimentado pelos dados
reais gravados no banco.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone

MESES = [
    'jan', 'fev', 'mar', 'abr', 'mai', 'jun',
    'jul', 'ago', 'set', 'out', 'nov', 'dez',
]

STATUS_LABEL = {
    'pendente':              'Aguardando aprovação',
    'aprovado':              'Aprovado / Liberado',
    'aguardando_devolucao':  'Aguard. confirmação devolução',
    'devolvido':             'Devolvido',
    'cancelado':             'Cancelado',
    'recusado':              'Recusado',
}

VERDE = '00CE7C'
AZUL  = '001A70'


def _fmt_data_longa(d):
    """Formata como '27 de jul. de 2026' — igual ao padrão usado na planilha."""
    if not d:
        return ''
    return f'{d.day:02d} de {MESES[d.month - 1]}. de {d.year}'


def _fmt_data_hora(dt):
    dt = timezone.localtime(dt)
    return dt.strftime('%d/%m/%Y às %H:%M:%S')


def gerar_relatorio_xlsx(pedidos_qs, periodo_label, usuario_nome):
    wb = Workbook()

    # ── Aba Resumo ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumo'

    total       = pedidos_qs.count()
    pendentes   = pedidos_qs.filter(status='pendente').count()
    em_posse    = pedidos_qs.filter(status__in=['aprovado', 'aguardando_devolucao']).count()
    devolvidos  = pedidos_qs.filter(status='devolvido').count()
    cancelados  = pedidos_qs.filter(status__in=['cancelado', 'recusado']).count()
    ocorrencias = pedidos_qs.exclude(ocorrencia__isnull=True).exclude(ocorrencia={}).count()
    taxa_ocorrencia = (ocorrencias / total * 100) if total else 0

    ws['A1'] = 'Controle Dressing · MRO — Relatório de Empréstimos'
    ws['A1'].font = Font(bold=True, size=13, color=AZUL)
    ws['A2'] = 'Período';     ws['B2'] = periodo_label
    ws['A3'] = 'Gerado em';   ws['B3'] = _fmt_data_hora(timezone.now())
    ws['A4'] = 'Responsável'; ws['B4'] = usuario_nome or ''

    ws['A6'] = 'Indicador'; ws['B6'] = 'Valor'
    for cell in ('A6', 'B6'):
        ws[cell].font = Font(bold=True)
        ws[cell].fill = PatternFill('solid', fgColor='E6EAF5')

    indicadores = [
        ('Total de pedidos', total),
        ('Aguardando aprovação', pendentes),
        ('Em posse do solicitante', em_posse),
        ('Devolvidos', devolvidos),
        ('Cancelados', cancelados),
        ('Ocorrências', ocorrencias),
        ('Taxa de ocorrência', f'{taxa_ocorrencia:.1f}%'),
    ]
    for i, (label, valor) in enumerate(indicadores, start=7):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = valor

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 36

    # ── Aba Pedidos ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Pedidos')
    headers = [
        'DATA_SOLICITAÇÃO', 'COORDENAÇÃO', 'ARMAZEM ORIGEM', 'CONTEÚDO - ORIGEM',
        'AUTORIZADO', 'LOCAL_ENTREGA', 'CONTEÚDO - DESTINO', 'SOLICITANTE',
        'PEDIDO', 'QUANTIDADE', 'MATERIAL', 'DATA DE RETORNO', 'NAO VOLTOU',
        'RETORNO', 'CÓDIGO', 'STATUS', 'OCORRÊNCIA (detalhe)',
    ]
    ws2.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=AZUL)
        cell.alignment = Alignment(horizontal='left')

    for p in pedidos_qs.order_by('-criado_em'):
        materiais = p.materiais if isinstance(p.materiais, list) else []
        ocorrencia_txt = ''
        if p.ocorrencia and isinstance(p.ocorrencia, dict):
            ocorrencia_txt = ' — '.join(
                filter(None, [p.ocorrencia.get('tipo'), p.ocorrencia.get('descricao')])
            )

        criado_local = timezone.localtime(p.criado_em) if p.criado_em else None
        nao_voltou = 'AINDA NÃO' if p.status in ('aprovado', 'aguardando_devolucao') and p.dev_iso else ''
        retorno    = 'RETORNO' if p.status == 'devolvido' else ''

        ws2.append([
            _fmt_data_longa(criado_local),
            p.concedente_nome if p.mg_concedente else '',
            p.mg_concedente or '',
            p.produto_concedente or '',
            p.concedente_nome or '',
            p.mg_solicitante or '',
            p.produto or '',
            p.solicitante_nome or '',
            p.numero_pedido or '',
            len(materiais) if materiais else 1,
            ' , '.join(materiais),
            _fmt_data_longa(p.dev_iso),
            nao_voltou,
            retorno,
            p.codigo or '',
            STATUS_LABEL.get(p.status, p.status),
            ocorrencia_txt,
        ])

    larguras = [16, 16, 14, 22, 16, 14, 22, 16, 10, 10, 40, 16, 10, 10, 14, 24, 26]
    for i, w in enumerate(larguras, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    return wb